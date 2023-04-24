import sqlite3
from openpyxl import Workbook, load_workbook

#imports xlsx_to_db.py from the same folder in order to create db file from sales.xlsx
#+++++++++++++++++++++++++++++++++
import xlsx_to_db

xlsx_to_db.main()
#=================================

filepath = "report.xlsx"
storages_like = ['Ташкентская', 'Рыскулова', 'КарТаун', 'Ломова', 'Естая']
#БРЕНД = Voltman Tubor
brands = ['БРЕНД', 'ENERGIZER', 'МУТЛУ', 'HYUNDAI', 'ТУБОР', 'Clarios', 'VARTA', 'КОРЕЯ', 'УКРАИНА', 'КИТАЙ', 'БАРС', 'КАЗАХСТАН']

#establishes connection to database
def connect_to_db(db_file):
    conn = 0
    try:
        conn = sqlite3.connect(db_file)
    except Error as e:
        print(e)

    return conn

def get_cities(conn):
    cur = conn.cursor()

    cur.execute('SELECT DISTINCT City FROM(SELECT DISTINCT City FROM current UNION ALL SELECT DISTINCT City FROM previous)')
    city_unedited = cur.fetchall()

    city_names = []
    for i in city_unedited:
        city_names.append(i[0])

    return city_names

def get_vehicle_types(conn):
    cur = conn.cursor()

    cur.execute("SELECT DISTINCT Vehicle FROM (SELECT DISTINCT Vehicle FROM current UNION ALL SELECT DISTINCT Vehicle FROM previous)")
    vehicle_types_unedited = cur.fetchall()

    vehicle_types = []
    for i in vehicle_types_unedited:
        vehicle_types.append(i[0])
    
    return vehicle_types

def get_storages(conn):
    cur = conn.cursor()

    cur.execute("SELECT DISTINCT Storage FROM current")
    current = cur.fetchall()
    
    cur.execute("SELECT DISTINCT Storage FROM previous")
    previous = cur.fetchall()

    storage_names = []

    for i in current:
        if i[0] not in storage_names:
            storage_names.append(i[0])

    for i in previous:
        if i[0] not in storage_names:
            storage_names.append(i[0])

    storage_names.sort()

    return storage_names

#create empty xlsx workbook
def create_workbook():
    wb = Workbook()
    wb.save(filepath)

#calculates discount sales total and by storage and adds them to sheet in xlsx file
def discount_total(conn):
    cur = conn.cursor()

    sum_current = 0
    sum_previous = 0
    storage_sales = {}
    discount = 'дисконт'
    for i in storages_like:
        sales = []
        cur.execute(f"SELECT SUM(Quantity) FROM current WHERE Storage LIKE '%{i}%' AND Storage LIKE '%{discount}%'")
        current = cur.fetchall()[0][0]
        if current is not None:
            sales.append(int(current))
        else:
            sales.append(0)
        cur.execute(f"SELECT SUM(Quantity) FROM previous WHERE Storage LIKE '%{i}%' AND Storage LIKE '%{discount}%'")
        previous = cur.fetchall()[0][0]
        if previous is not None:
            sales.append(int(previous))
        else:
            sales.append(0)

        storage_sales[f"{i}"] = sales
        
        sum_current += sales[0]
        sum_previous += sales[1]

    storage_sales['Total'] = [sum_current, sum_previous]

    wb = load_workbook(filepath)

    ws = wb.create_sheet('Дисконт')

    data = [('Склад', 'current', 'previous')]
    
    for i in storage_sales:
        dt = (i, storage_sales[i][0], storage_sales[i][1])
        data.append(dt)

    for row in data:
        ws.append(row)

    wb.save(filepath)

#calculates total sales and adds them to xlsx file
def total_sales(conn):
    cur = conn.cursor()

    sum_current = 0
    sum_previous = 0
    shop_sales = {}
    for i in storages_like:
        sales = []
        cur.execute(f"SELECT SUM(Quantity) FROM current WHERE Storage LIKE '%{i}%'")
        current = cur.fetchall()[0][0]
        sales.append(current)
        if current != None:
            sum_current += current
        cur.execute(f"SELECT SUM(Quantity) FROM previous WHERE Storage LIKE '%{i}%'")
        previous = cur.fetchall()[0][0]
        sales.append(previous)
        if previous != None:
            sum_previous += previous

        shop_sales[i] = sales

    shop_sales['Total'] = [sum_current, sum_previous]

    wb = load_workbook(filepath)
    ws = wb.create_sheet('Продажи')
    data = [('Склад', 'current', 'previous', 'Динамика')]

    for i in shop_sales:
        if shop_sales[i][0] != None and shop_sales[i][1] != None:
            dt = (i, shop_sales[i][0], shop_sales[i][1], f'{round((shop_sales[i][0]-shop_sales[i][1])/shop_sales[i][1]*100, 1)}%')
        else:
            dt = (i, shop_sales[i][0], shop_sales[i][1], 0)
        data.append(dt)

    

    for row in data:
        ws.append(row)
    
    wb.save(filepath)

#calculates avarage price by storage and city and adds them to xlsx file
def avarage(conn):
    cities = get_cities(conn)

    cur = conn.cursor()

    average = {}
    for i in cities:
        for j in storages_like:
            cur.execute(f"SELECT (SUM(Price)/SUM(Quantity)) FROM current WHERE Storage LIKE '%{j}%' AND City = '{i}'")
            current = cur.fetchall()[0][0]
            cur.execute(f"SELECT (SUM(Price)/SUM(Quantity)) FROM previous WHERE Storage LIKE '%{j}%' AND City = '{i}'")
            previous = cur.fetchall()[0][0]
            if current == None:
                current = 0
            if previous == None:
                previous = 0
            average[f'{j} {i}'] = [round(current, 1), round(previous, 1)]
        cur.execute(f"SELECT (SUM(Price)/SUM(Quantity)) FROM current WHERE City = '{i}'")
        current_city = cur.fetchall()[0][0]
        if current_city == None:
            current_city = 0
        cur.execute(f"SELECT (SUM(Price)/SUM(Quantity)) FROM previous WHERE City = '{i}'")
        previous_city = cur.fetchall()[0][0]
        if previous_city == None:
            previous_city =0
        average[i] = [round(current_city, 1), round(previous_city, 1)]

    wb = load_workbook(filepath)
    ws = wb.create_sheet('Средний')
    data = [('Склад', 'current', 'previous')]

    for i in average:
        data.append((i, average[i][0], average[i][1]))

    for row in data:
        ws.append(row)

    wb.save(filepath)

def sales_by_vehicle(conn):
    cur = conn.cursor()

    vehicles = get_vehicle_types(conn)

    by_storage = {}
    for i in storages_like:
        storage_data = []
        for j in vehicles:
            cur.execute(f"SELECT SUM(Quantity), SUM(Price) FROM current WHERE Storage LIKE '%{i}%' AND Vehicle = '{j}'")
            values = cur.fetchall()[0]
            count = values[0]
            price = values[1]
            if count == None:
                count = 0
                price = 0
            data = (count, price)
            storage_data.append(data)
        by_storage[i] = storage_data
    
    storage_data = []
    for i in vehicles:
        cur.execute(f"SELECT SUM(Quantity), SUM(Price) FROM current WHERE Vehicle = '{i}'")
        values = cur.fetchall()[0]
        count = values[0]
        price = values[1]
        if count == None:
                count = 0
                price = 0
        data = (count, price)
        storage_data.append(data)
    by_storage['Total'] = storage_data

    wb = load_workbook(filepath)

    ws = wb.create_sheet('Тип транспорта')

    data = [('Количество', vehicles[0], vehicles[1], vehicles[2])]

    for i in by_storage:
        data.append((i, by_storage[i][0][0], by_storage[i][1][0], by_storage[i][2][0]))

    for row in data:
        ws.append(row)

    ws.append(("","","",""))

    data = [('Доля от количества', vehicles[0], vehicles[1], vehicles[2])]

    for i in by_storage:
        data.append((i, f"{round(by_storage[i][0][0]/(by_storage[i][0][0] + by_storage[i][1][0] + by_storage[i][2][0])*100, 1)}%", f"{round(by_storage[i][1][0]/(by_storage[i][0][0] + by_storage[i][1][0] + by_storage[i][2][0])*100, 1)}%", f"{round(by_storage[i][2][0]/(by_storage[i][0][0] + by_storage[i][1][0] + by_storage[i][2][0])*100, 1)}%"))

    for row in data:
        ws.append(row)

    ws.append(("","","",""))

    data = [('Доля от суммы', vehicles[0], vehicles[1], vehicles[2])]

    for i in by_storage:
        data.append((i, f"{round(by_storage[i][0][1]/(by_storage[i][0][1] + by_storage[i][1][1] + by_storage[i][2][1])*100, 1)}%", f"{round(by_storage[i][1][1]/(by_storage[i][0][1] + by_storage[i][1][1] + by_storage[i][2][1])*100, 1)}%", f"{round(by_storage[i][2][1]/(by_storage[i][0][1] + by_storage[i][1][1] + by_storage[i][2][1])*100, 1)}%"))

    for row in data:
        ws.append(row)

    wb.save(filepath)

#calculatioins by brand and output to xlsx file
def brand(conn):
    cur = conn.cursor()

    by_brand = {}
    for i in storages_like:
        brand_data = []
        for j in brands:
            cur.execute(f"SELECT SUM(Quantity), SUM(Price) FROM current WHERE Storage LIKE '%{i}%' AND Brand LIKE '%{j}%'")
            values = cur.fetchall()[0]
            count = values[0]
            price = values[1]
            data = (count, price)
            brand_data.append(data)
        by_brand[i] = brand_data

    wb = load_workbook(filepath)
    ws = wb.create_sheet("Бренд")

    #for quantitiy numbers by brands
    #++++++++++++++++++++++++++++++++++++++++++++++++++++
    data = brands.copy()    #copies brands list for iteration
    data.insert(0, "Бренды")    #adds first column to the list, required for sheet append
    data = [tuple(data)]    #data is converted to list of tuples containing data, required for sheet append
    
    for i in by_brand:
        tup = []
        for index, j in enumerate(brands):
            tup.append(by_brand[i][index][0])
        tup.insert(0, i)
        data.append(tuple(tup))

    total_by_brands = []
    for i in brands:
        cur.execute(f"SELECT SUM(Quantity) FROM current WHERE Brand LIKE '%{i}%'")
        total_by_brands.append(cur.fetchall()[0][0])
    total_by_brands.insert(0, "Total")

    data.append(tuple(total_by_brands))

    for row in data:
        ws.append(row)
    #====================================================

    ws.append(["", "", "", "", "", "", "", "", "", "", ""])

    #for quantity precentage values of brands
    #++++++++++++++++++++++++++++++++++++++++++++++++++++
    data = brands.copy()
    data.insert(0, "Бренды % колличество")
    data = [tuple(data)]

    for i in by_brand:
        tup = []
        for index, j in enumerate(brands):
            cur.execute(f"SELECT SUM(Quantity) FROM current WHERE Storage LIKE '%{i}%'")
            total_quantity_by_storage = cur.fetchall()[0][0]
            if by_brand[i][index][0] == None:
                tup.append(None)
                continue
            tup.append(f"{round(by_brand[i][index][0]/total_quantity_by_storage*100, 1)}%")
        tup.insert(0, i)
        data.append(tuple(tup))

    cur.execute(f"SELECT SUM(Quantity) FROM current")
    quantity_total = cur.fetchall()[0][0]
    data_by_brands = []
    for i in brands:
        cur.execute(f"SELECT SUM(Quantity) FROM current WHERE Brand LIKE '%{i}%'")
        quantity_by_brand = cur.fetchall()[0][0]
        if quantity_by_brand == None:
            data_by_brands.append(None)
            continue
        data_by_brands.append(f"{round(quantity_by_brand/quantity_total*100, 1)}%")
    data_by_brands.insert(0, "Total")

    data.append(data_by_brands)

    for row in data:
        ws.append(row)
    #====================================================

    ws.append(["", "", "", "", "", "", "", "", "", "", ""])

    #for price precentage values of brands
    #++++++++++++++++++++++++++++++++++++++++++++++++++++
    data = brands.copy()
    data.insert(0, "Бренды % сумма")
    data = [tuple(data)]

    for i in by_brand:
        tup = []
        for index, j in enumerate(brands):
            cur.execute(f"SELECT SUM(Price) FROM current WHERE Storage LIKE '%{i}%'")
            total_price_by_storage = cur.fetchall()[0][0]
            if by_brand[i][index][1] == None:
                tup.append(None)
                continue
            tup.append(f"{round(by_brand[i][index][1]/total_price_by_storage*100, 1)}%")
        tup.insert(0, i)
        data.append(tuple(tup))

    cur.execute(f"SELECT SUM(Price) FROM current")
    price_total = cur.fetchall()[0][0]
    data_by_brands = []
    for i in brands:
        cur.execute(f"SELECT SUM(Price) FROM current WHERE Brand LIKE '%{i}%'")
        price_by_brand = cur.fetchall()[0][0]
        if price_by_brand == None:
            data_by_brands.append(None)
            continue
        data_by_brands.append(f"{round(price_by_brand/price_total*100, 1)}%")
    data_by_brands.insert(0, "Total")

    data.append(data_by_brands)

    for row in data:
        ws.append(row)
    #====================================================

    wb.save(filepath)

#calculations by battery size
def bodytype(conn):
    cur = conn.cursor()

    cur.execute(f"SELECT DISTINCT Bodytype FROM current")
    bodytype_list_unedited = cur.fetchall()
    bodytype_list = []
    for i in bodytype_list_unedited:
        bodytype_list.append(i[0])
    
    cities = get_cities(conn)


    data = []

    for city in cities:
        data.append(("Моноблок", "Количество", "Сумма","Доля от количества", "Доля от суммы"))
        cur.execute(f"SELECT SUM(Quantity), SUM(Price) FROM current WHERE City = '{city}'")
        city_total = cur.fetchall()
        city_quantity = city_total[0][0]
        if city_quantity == None:
            continue
        city_price = city_total[0][1]
        city_total = (city, city_quantity, city_price, "100.0%", "100.0%")
        for bodytype in bodytype_list:
            cur.execute(f"SELECT SUM(Quantity), SUM(Price) FROM current WHERE City = '{city}' AND Bodytype = '{bodytype}'")
            batterytype = cur.fetchall()
            batterytype_quantity = batterytype[0][0]
            batterytype_price = batterytype[0][1]
            if batterytype_quantity == None:
                data.append((bodytype, batterytype_quantity, batterytype_price, None, None))
                continue
            data.append((bodytype, batterytype_quantity, batterytype_price, f"{round(batterytype_quantity/city_quantity*100, 1)}%", f"{round(batterytype_price/city_price*100, 1)}%"))
        data.append(city_total)
        data.append(("", "", "", "", "", ))

    wb = load_workbook(filepath)
    ws = wb.create_sheet("Моноблок")

    for row in data:
        ws.append(row)
    
    wb.save(filepath)

def manage_sales(conn):
    cur = conn.cursor()

    cur.execute(f"SELECT DISTINCT Manager FROM current ")

    manager_list_unedited = cur.fetchall()
    manager_list = []
    for i in manager_list_unedited:
        manager_list.append(i[0])

    manager_sales = [("Менеджер", "Количество", "Сумма", "Средняя стоимость", "Доля продаж в магазине (количество)", "Доля продаж в магазине (сумма)")]
    for i in storages_like:
        for j in manager_list:
            cur.execute(f"SELECT SUM(Quantity), SUM(Price) FROM current WHERE Manager = '{j}' AND Storage LIKE '%{i}%'")
            sales = cur.fetchall()[0]
            quantity = sales[0]
            price = sales[1]
            if quantity == None:
                continue
            avg = price/(quantity if (quantity != 0) else 1)
            cur.execute(f"SELECT SUM(Quantity), SUM(Price) FROM current WHERE Storage LIKE '%{i}%'")
            sales = cur.fetchall()[0]
            storage_quantity = sales[0]
            storage_price = sales[1]
            percentage_quantity = quantity/storage_quantity*100
            percentage_price = price/storage_price*100
            manager_sales.append((f"{j} ({i})",quantity, price, round(avg, 0), f"{round(percentage_quantity, 1)}%", f"{round(percentage_price, 1)}%"))
        
    wb = load_workbook(filepath)
    ws = wb.create_sheet("По менеджерам")

    for row in manager_sales:
        ws.append(row)

    wb.save(filepath)

def manager_by_vehicle(conn):
    cur = conn.cursor()

    vehicles = get_vehicle_types(conn)

    total_by_vehicle = []
    for i in vehicles:
        cur.execute(f"SELECT SUM(Quantity), SUM(Price) FROM current WHERE Vehicle = '{i}'")
        sales = cur.fetchall()[0]
        count = sales[0]
        price = sales[1]
        if count == 0:
            count = 0
            price = 0
        total_by_vehicle.append((count, price))

    cur.execute(f"SELECT DISTINCT Manager FROM current ")

    manager_list_unedited = cur.fetchall()
    manager_list = []
    for i in manager_list_unedited:
        manager_list.append(i[0])

    by_manager = {}
    for i in manager_list:
        by_vehicle = []
        for j in vehicles:
            cur.execute(f"SELECT SUM(Quantity), SUM(Price) FROM current WHERE Manager = '{i}' AND Vehicle = '{j}'")
            sales = cur.fetchall()[0]
            count = sales[0]
            price = sales[1]
            by_vehicle.append((count, price))
        by_manager[i] = by_vehicle

    wb = load_workbook(filepath)
    ws = wb.create_sheet("Менеджер-транспорт")

    #header of table in xlsx for quantity
    #+++++++++++++++++++++++
    data = []
    for i in vehicles:
        data.append(i)
    data.insert(0, "Менеджер-Количество")
    data = [tuple(data)]
    #=======================

    #for quantity by vehicle
    #+++++++++++++++++++++++
    for i in by_manager:
        by_vehicle = []
        for index, j in enumerate(vehicles):
            by_vehicle.append(by_manager[i][index][0])
        by_vehicle.insert(0, i)
        data.append(tuple(by_vehicle))
    
    for row in data:
        ws.append(row)
    #=======================

    ws.append(("", "", "", ""))

    #header of table in xlsx for quantity-percentages
    #+++++++++++++++++++++++
    data = []
    for i in vehicles:
        data.append(i)
    data.insert(0, "Менеджер-количество-проценты в разрезе собственных продаж")
    data = [tuple(data)]
    #=======================

    #for quantity percentage by vehicle
    #+++++++++++++++++++++++
    quantity_by_manager = {}
    for i in manager_list:
        cur.execute(f"SELECT SUM(Quantity) FROM current WHERE Manager = '{i}'")
        quantity_by_manager[i] = cur.fetchall()[0][0]

    for i in by_manager:
        by_vehicle = []
        for index, j in enumerate(vehicles):
            if by_manager[i][index][0] == None:
                by_vehicle.append(None)
                continue
            by_vehicle.append(f"{round(by_manager[i][index][0]/quantity_by_manager[i]*100, 1)}%")
        by_vehicle.insert(0, i)
        data.append(tuple(by_vehicle))
    
    for row in data:
        ws.append(row)
    #=======================

    ws.append(("", "", "", ""))

    wb.save(filepath)

def manager_brand(conn):
    cur = conn.cursor()

    cur.execute(f"SELECT DISTINCT Manager FROM current ")

    manager_list_unedited = cur.fetchall()
    manager_list = []
    for i in manager_list_unedited:
        manager_list.append(i[0])

    wb = load_workbook(filepath)
    ws = wb.create_sheet("Менеджер-бренд")

    #for brand quantities
    #+++++++++++++++++++++++++++++++++++++++++
    data = brands.copy()
    data.insert(0, "Менеджер")
    data = [tuple(data)]
    
    for i in manager_list:
        by_brands = []
        for j in brands:
            cur.execute(f"SELECT SUM(Quantity) FROM current WHERE Manager = '{i}' AND Brand LIKE '%{j}%'")
            quantity = cur.fetchall()[0][0]
            by_brands.append(quantity)
        by_brands.insert(0, i)
        data.append(tuple(by_brands))

    for row in data:
        ws.append(row)
    #=============================================
    

    #for brand quantity percentages
    #+++++++++++++++++++++++++++++++++++++++++
    data = brands.copy()
    data.insert(0, "Менеджер")
    data = [tuple(data)]
    
    for i in manager_list:
        by_brands = []
        cur.execute(f"SELECT SUM(Quantity) FROM current WHERE Manager = '{i}'")
        quantity_total = cur.fetchall()[0][0]
        if quantity_total == None:
            continue
        for j in brands:
            cur.execute(f"SELECT SUM(Quantity) FROM current WHERE Manager = '{i}' AND Brand LIKE '%{j}%'")
            quantity = cur.fetchall()[0][0]
            if quantity == None:
                by_brands.append(None)
                continue
            by_brands.append(f"{round(quantity/quantity_total*100, 1)}%")
        by_brands.insert(0, i)
        data.append(tuple(by_brands))

    for row in data:
        ws.append(row)
    #=============================================

    wb.save(filepath)

def main():
    database = "sales.db"
    conn = connect_to_db(database)

    with conn:
        get_cities(conn)

        get_storages(conn)

        get_vehicle_types(conn)

        create_workbook()

        discount_total(conn)

        total_sales(conn)

        avarage(conn)

        sales_by_vehicle(conn)

        brand(conn)

        bodytype(conn)

        manage_sales(conn)

        manager_by_vehicle(conn)

        manager_brand(conn)

if __name__ == "__main__":
    main()